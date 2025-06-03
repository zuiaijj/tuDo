import sys
import subprocess

# 检查并自动安装openpyxl
try:
    import openpyxl # type: ignore
except ImportError:
    print("未检测到openpyxl，正在自动安装...")
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    print("openpyxl 安装完成，请重新运行本脚本。")
    sys.exit(0)
from openpyxl import load_workbook # type: ignore

if len(sys.argv) < 2:
    print("请传入xlsx文件路径")
    sys.exit(1)

file_path = sys.argv[1]
target_sheet = "Tudo客户端"

wb = load_workbook(file_path)
sheetnames = wb.sheetnames

print(f"获取到的sheetnames: {sheetnames}")

if target_sheet in sheetnames:
    # 将目标sheet移到第一个
    wb._sheets.insert(0, wb._sheets.pop(sheetnames.index(target_sheet)))
    ws = wb[target_sheet]
    # 删除前4列（A-D）
    ws.delete_cols(1, 2)
    # 修改第一行内容
    header = [
        "tip_title",
        "Chinese(intl_zh)",
        "English(intl_en)",
        "TaiWan(intl_zh_TW)",
    ]
    for idx, value in enumerate(header, start=1):
        ws.cell(row=1, column=idx, value=value)
    wb.save(file_path)
    print(f"已将 {target_sheet} 移动到第一个sheet，并处理表头和列")
else:
    print(f"未找到sheet: {target_sheet}，无需调整") 